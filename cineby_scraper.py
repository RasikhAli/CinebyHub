"""
==========================================================================
  Cineby.gd Content Scraper - Powered by TMDB API
==========================================================================
  Fetches ALL movies, TV shows, anime, and live/broadcast channels
  from The Movie Database (TMDB) — the same data source used by cineby.gd
  and Vidking Player (which uses /embed/movie/{tmdbId} and /embed/tv/{tmdbId}).

  Usage:
    1. Set your credentials in the CONFIG section below, OR
       set environment variables:  TMDB_API_KEY  /  TMDB_READ_TOKEN
    2. Run: python cineby_scraper.py
    3. Output: cineby_content.xlsx  (in the same directory)
==========================================================================
"""

import os
import sys
import time
import json
import requests
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG — paste your credentials here, OR let the script ask you at runtime
# ─────────────────────────────────────────────────────────────────────────────
TMDB_API_KEY    = os.getenv("TMDB_API_KEY", "")
TMDB_READ_TOKEN = os.getenv("TMDB_READ_TOKEN", "")

# ── Auto-load from .env file if present ───────────────────────────────────────
def _load_env():
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, _, v = line.partition("=")
                    os.environ.setdefault(k.strip(), v.strip())

_load_env()
# Already set above via os.getenv, but keeping the fallback logic safe
TMDB_API_KEY    = TMDB_API_KEY    or os.getenv("TMDB_API_KEY", "")
TMDB_READ_TOKEN = TMDB_READ_TOKEN or os.getenv("TMDB_READ_TOKEN", "")

# How many pages to fetch per endpoint (TMDB caps at 500 pages, 20 items/page)
# Set to None to fetch ALL available pages (may take a while)
MAX_PAGES_MOVIES    = None   # None = fetch ALL  |  e.g. 50 = first 1000 movies
MAX_PAGES_TV        = None   # None = fetch ALL
MAX_PAGES_ANIME     = None   # None = fetch ALL
MAX_PAGES_CHANNELS  = 10     # TMDB networks list is small, 10 is usually enough

OUTPUT_FILE         = "cineby_content.xlsx"
REQUEST_DELAY       = 0.25   # seconds between API calls (respect rate limits)

# ── Incremental / Resume Mode ─────────────────────────────────────────────────
# If True  → on re-run, load existing Excel data and only add NEW items.
#            Stops fetching a page-list as soon as it sees IDs already saved.
# If False → always start fresh and overwrite the output file completely.
INCREMENTAL_UPDATE = True
# ─────────────────────────────────────────────────────────────────────────────

BASE_URL = "https://api.themoviedb.org/3"
IMAGE_BASE = "https://image.tmdb.org/t/p/w185"
CINEBY_BASE = "https://www.cineby.gd"
VIDKING_BASE = "https://www.vidking.net"

# TMDB Genre ID → Name mapping
GENRE_MAP = {
    28: "Action", 12: "Adventure", 16: "Animation", 35: "Comedy",
    80: "Crime", 99: "Documentary", 18: "Drama", 10751: "Family",
    14: "Fantasy", 36: "History", 27: "Horror", 10402: "Music",
    9648: "Mystery", 10749: "Romance", 878: "Sci-Fi", 10770: "TV Movie",
    53: "Thriller", 10752: "War", 37: "Western",
    # TV-specific
    10759: "Action & Adventure", 10762: "Kids", 10763: "News",
    10764: "Reality", 10765: "Sci-Fi & Fantasy", 10766: "Soap",
    10767: "Talk", 10768: "War & Politics",
}

# ─────────────────────────────────────────────────────────────────────────────

def check_dependencies():
    missing = []
    for pkg in ["requests", "openpyxl", "tqdm"]:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"[!] Missing packages: {', '.join(missing)}")
        print(f"    Install with: pip install {' '.join(missing)}")
        sys.exit(1)

check_dependencies()

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from tqdm import tqdm


# ─────────────────────────────────────────────────────────────────────────────
#  API CLIENT
# ─────────────────────────────────────────────────────────────────────────────

class TMDBClient:
    def __init__(self):
        token = TMDB_READ_TOKEN
        key   = TMDB_API_KEY

        # Prefer Bearer token (v4 auth), fall back to API key (v3)
        if token:
            self.headers = {
                "Authorization": f"Bearer {token}",
                "accept": "application/json",
            }
            self.params = {}
            print("[✓] Using Bearer token (Read Access Token) authentication")
        elif key:
            self.headers = {"accept": "application/json"}
            self.params = {"api_key": key}
            print("[✓] Using API key authentication")
        else:
            print("[!] No TMDB credentials found in script config or .env file.")
            print("    Get your free credentials at: https://developer.themoviedb.org/")
            print()
            print("    [1] API Key (v3)")
            print("    [2] Read Access Token / Bearer Token (v4, recommended)")
            choice = input("    Your choice [1/2]: ").strip()
            if choice == "1":
                key = input("    Paste your API Key: ").strip()
                self.headers = {"accept": "application/json"}
                self.params = {"api_key": key}
            else:
                token = input("    Paste your Read Access Token: ").strip()
                self.headers = {
                    "Authorization": f"Bearer {token}",
                    "accept": "application/json",
                }
                self.params = {}
            print("[✓] Credentials accepted")

        self.session = requests.Session()
        self.session.headers.update(self.headers)

    def get(self, endpoint, extra_params=None):
        params = {**self.params}
        if extra_params:
            params.update(extra_params)
        url = f"{BASE_URL}{endpoint}"
        for attempt in range(3):
            try:
                resp = self.session.get(url, params=params, timeout=15)
                if resp.status_code == 429:  # rate limited
                    wait = int(resp.headers.get("Retry-After", 5))
                    print(f"\n[!] Rate limited — waiting {wait}s...")
                    time.sleep(wait)
                    continue
                resp.raise_for_status()
                return resp.json()
            except requests.RequestException as e:
                if attempt == 2:
                    print(f"\n[!] Request failed after 3 attempts: {e}")
                    return None
                time.sleep(2 ** attempt)
        return None

    def fetch_all_pages(self, endpoint, params=None, max_pages=None, desc="Fetching"):
        """Fetch all pages from a paginated TMDB endpoint."""
        results = []
        page = 1
        total_pages = 1

        with tqdm(desc=desc, unit=" pages", dynamic_ncols=True) as pbar:
            while page <= total_pages:
                if max_pages and page > max_pages:
                    break
                data = self.get(endpoint, {**(params or {}), "page": page})
                if not data:
                    break

                page_results = data.get("results", [])
                results.extend(page_results)

                total_pages = min(
                    data.get("total_pages", 1),
                    500  # TMDB hard cap per query. 
                         # We bypass this by 'slicing' (yearly/regional queries).
                )
                pbar.total = min(total_pages, max_pages or total_pages)
                pbar.set_postfix({"items": len(results), "total_pages": total_pages})
                pbar.update(1)

                page += 1
                time.sleep(REQUEST_DELAY)

        return results


# ─────────────────────────────────────────────────────────────────────────────
#  INCREMENTAL HELPERS — load what's already saved in the Excel file
# ─────────────────────────────────────────────────────────────────────────────

def load_existing_rows(output_path, sheet_name, id_col="TMDB ID"):
    """
    Read all rows from a sheet in an existing Excel file.
    Returns (list_of_dicts, set_of_ids).
    Returns ([], set()) if the file or sheet doesn't exist.
    """
    if not os.path.exists(output_path):
        return [], set()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return [], set()
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        data = []
        ids  = set()
        id_idx = headers.index(id_col) if id_col in headers else None
        for row in rows_iter:
            row_dict = dict(zip(headers, row))
            # Restore None → "" for cleanliness
            row_dict = {k: (v if v is not None else "") for k, v in row_dict.items()}
            data.append(row_dict)
            if id_idx is not None and row[id_idx] is not None:
                ids.add(row[id_idx])
        wb.close()
        return data, ids
    except Exception as e:
        print(f"  [!] Could not read existing sheet '{sheet_name}': {e}")
        return [], set()


def fetch_incremental(client, endpoint, params, existing_ids, max_pages, desc):
    """
    Like fetch_all_pages but stops early once a full page of results
    contains only IDs that already exist in existing_ids.
    This means the dataset is up-to-date from that page onward.
    """
    results = []
    page = 1
    total_pages = 1

    with tqdm(desc=desc, unit=" pages", dynamic_ncols=True) as pbar:
        while page <= total_pages:
            if max_pages and page > max_pages:
                break
            data = client.get(endpoint, {**(params or {}), "page": page})
            if not data:
                break

            page_results = data.get("results", [])
            total_pages = min(data.get("total_pages", 1), 500)
            pbar.total = min(total_pages, max_pages or total_pages)

            new_on_page = [r for r in page_results if r.get("id") not in existing_ids]
            results.extend(new_on_page)

            pbar.set_postfix({
                "new": len(results),
                "page_new": len(new_on_page),
                "total_pages": total_pages,
            })
            pbar.update(1)

            # If the entire page had no new items → TMDB is sorted by
            # popularity/date so older items follow — safe to stop here.
            if len(new_on_page) == 0 and page > 1:
                pbar.set_postfix({"status": "up-to-date ✓", "new": len(results)})
                break

            page += 1
            time.sleep(REQUEST_DELAY)

    return results


# ─────────────────────────────────────────────────────────────────────────────
#  DATA FETCHERS
# ─────────────────────────────────────────────────────────────────────────────

def dedupe(items, key="id"):
    seen = set()
    out = []
    for item in items:
        v = item.get(key)
        if v and v not in seen:
            seen.add(v)
            out.append(item)
    return out


def get_genre_names(genre_ids):
    return ", ".join(GENRE_MAP.get(gid, str(gid)) for gid in (genre_ids or []))


def normalize_movie(item):
    tmdb_id = item.get("id", "")
    return {
        "TMDB ID":         tmdb_id,
        "Title":           item.get("title", item.get("name", "")),
        "Original Title":  item.get("original_title", item.get("original_name", "")),
        "Overview":        item.get("overview", ""),
        "Release Date":    item.get("release_date", ""),
        "Rating (TMDB)":   round(item.get("vote_average", 0), 1),
        "Vote Count":      item.get("vote_count", 0),
        "Popularity":      round(item.get("popularity", 0), 2),
        "Language":        item.get("original_language", ""),
        "Genres":          get_genre_names(item.get("genre_ids", [])),
        "Poster":          f"{IMAGE_BASE}{item.get('poster_path', '')}" if item.get("poster_path") else "",
        "Backdrop":        f"https://image.tmdb.org/t/p/original{item.get('backdrop_path', '')}" if item.get("backdrop_path") else "",
        "Cineby URL":      f"{CINEBY_BASE}/movie/{tmdb_id}",
        "Vidking Embed":   f"{VIDKING_BASE}/embed/movie/{tmdb_id}",
        "Adult":           item.get("adult", False),
    }


def normalize_tv(item, is_anime=False):
    tmdb_id = item.get("id", "")
    return {
        "TMDB ID":         tmdb_id,
        "Title":           item.get("name", item.get("title", "")),
        "Original Title":  item.get("original_name", item.get("original_title", "")),
        "Overview":        item.get("overview", ""),
        "First Air Date":  item.get("first_air_date", ""),
        "Rating (TMDB)":   round(item.get("vote_average", 0), 1),
        "Vote Count":      item.get("vote_count", 0),
        "Popularity":      round(item.get("popularity", 0), 2),
        "Language":        item.get("original_language", ""),
        "Genres":          get_genre_names(item.get("genre_ids", [])),
        "Origin Country":  ", ".join(item.get("origin_country", [])),
        "Is Anime":        "Yes" if is_anime else "No",
        "Poster":          f"{IMAGE_BASE}{item.get('poster_path', '')}" if item.get("poster_path") else "",
        "Backdrop":        f"https://image.tmdb.org/t/p/original{item.get('backdrop_path', '')}" if item.get("backdrop_path") else "",
        "Cineby URL":      f"{CINEBY_BASE}/tv/{tmdb_id}",
        "Cineby Ep1 URL":  f"{CINEBY_BASE}/tv/{tmdb_id}/1/1",
        "Vidking Embed":   f"{VIDKING_BASE}/embed/tv/{tmdb_id}/1/1",
    }


def normalize_channel(item):
    return {
        "Network ID":      item.get("id", ""),
        "Name":            item.get("name", ""),
        "Country":         item.get("origin_country", ""),
        "Logo":            f"https://image.tmdb.org/t/p/w185{item.get('logo_path', '')}" if item.get("logo_path") else "",
        "Headquarters":    item.get("headquarters", ""),
        "Homepage":        item.get("homepage", ""),
        "TMDB Page":       f"https://www.themoviedb.org/network/{item.get('id', '')}",
    }


def fetch_networks(client, max_pages=10):
    """
    TMDB doesn't have a direct 'list all networks' endpoint.
    We discover them by querying popular TV shows and collecting
    unique networks. Also fetch the most popular TMDB networks directly.
    """
    print("\n📡  Fetching TV Networks / Channels...")
    networks = {}

    # Discover networks from popular TV shows
    tv_data = client.fetch_all_pages(
        "/discover/tv",
        params={"sort_by": "popularity.desc", "language": "en-US"},
        max_pages=max_pages,
        desc="  Scanning TV shows for networks"
    )

    network_ids = set()
    for show in tv_data:
        for nid in show.get("networks", []):
            if isinstance(nid, dict):
                network_ids.add(nid.get("id"))
            elif isinstance(nid, int):
                network_ids.add(nid)

    # Well-known streaming/broadcast network IDs to always include
    known_networks = [
        213, 49, 2739, 174, 453, 2552, 2625, 359, 19, 6,
        2, 1, 56, 182, 318, 67, 4, 34, 71, 16,
        # International
        289, 288, 4330, 1024, 510, 247, 2993, 5764, 337, 591,
    ]
    network_ids.update(known_networks)

    print(f"  → Fetching details for {len(network_ids)} unique networks...")
    detailed = []
    for nid in tqdm(network_ids, desc="  Network details", unit=" networks"):
        if nid is None:
            continue
        data = client.get(f"/network/{nid}")
        if data and data.get("id"):
            detailed.append({
                "Network ID":   data.get("id", ""),
                "Name":         data.get("name", ""),
                "Country":      data.get("origin_country", ""),
                "Logo":         f"https://image.tmdb.org/t/p/w185{data.get('logo_path', '')}" if data.get("logo_path") else "",
                "Headquarters": data.get("headquarters", ""),
                "Homepage":     data.get("homepage", ""),
                "TMDB Page":    f"https://www.themoviedb.org/network/{nid}",
            })
        time.sleep(REQUEST_DELAY)

    return detailed


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL WRITER
# ─────────────────────────────────────────────────────────────────────────────

# Color palette
COLOR_DARK_BG     = "141414"
COLOR_HEADER_MOV  = "E50914"   # Netflix red — Movies
COLOR_HEADER_TV   = "0dcaf0"   # Cineby cyan — TV Shows
COLOR_HEADER_ANI  = "9146FF"   # Twitch purple — Anime
COLOR_HEADER_CH   = "1DB954"   # Spotify green — Channels
COLOR_HEADER_SUM  = "F5A623"   # Gold — Summary
COLOR_ROW_ALT     = "1C1C1C"
COLOR_ROW_MAIN    = "111111"
COLOR_TEXT_LIGHT  = "FFFFFF"
COLOR_TEXT_DIM    = "AAAAAA"
COLOR_LINK        = "0dcaf0"


def make_header_style(hex_color, bold=True):
    fill = PatternFill("solid", fgColor=hex_color)
    font = Font(color=COLOR_TEXT_LIGHT, bold=bold, size=11, name="Segoe UI")
    align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    border = Border(
        bottom=Side(style="thin", color="000000"),
    )
    return fill, font, align, border


def apply_headers(ws, headers, color):
    fill, font, align, border = make_header_style(color)
    ws.row_dimensions[1].height = 32
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def style_data_row(ws, row_idx, num_cols, url_cols=None, url_col_indices=None):
    fill_color = COLOR_ROW_ALT if row_idx % 2 == 0 else COLOR_ROW_MAIN
    fill = PatternFill("solid", fgColor=fill_color)
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        is_url = url_col_indices and col_idx in url_col_indices
        cell.font = Font(
            color=COLOR_LINK if is_url else COLOR_TEXT_LIGHT,
            size=10,
            name="Segoe UI",
            underline="single" if is_url else None,
        )
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=(col_idx == 3),  # wrap Overview column
        )


def set_col_widths(ws, headers, widths=None):
    default_widths = {
        "TMDB ID": 10, "Network ID": 10, "Title": 35, "Original Title": 30,
        "Overview": 55, "Description": 55, "Release Date": 14,
        "First Air Date": 14, "Rating (TMDB)": 14, "Vote Count": 12,
        "Popularity": 12, "Language": 10, "Genres": 30, "Origin Country": 16,
        "Is Anime": 10, "Adult": 8, "Poster": 16, "Backdrop": 16,
        "Cineby URL": 40, "Cineby Ep1 URL": 42, "Vidking Embed": 42,
        "Name": 35, "Country": 10, "Logo": 16,
        "Headquarters": 25, "Homepage": 35, "TMDB Page": 40,
    }
    for col_idx, header in enumerate(headers, 1):
        width = (widths or {}).get(header, default_widths.get(header, 18))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_sheet(wb, sheet_name, tab_color, header_color, data_rows, headers,
                url_headers=None):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = tab_color
    ws.freeze_panes = "A2"

    # Dark background for entire sheet (cosmetic — fill blank area)
    ws.sheet_view.showGridLines = False
    apply_headers(ws, headers, header_color)
    set_col_widths(ws, headers)

    url_col_indices = set()
    if url_headers:
        for i, h in enumerate(headers, 1):
            if h in url_headers:
                url_col_indices.add(i)

    for row_idx, row_data in enumerate(data_rows, 2):
        ws.row_dimensions[row_idx].height = 22
        for col_idx, header in enumerate(headers, 1):
            val = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            is_url = col_idx in url_col_indices
            fill_color = COLOR_ROW_ALT if row_idx % 2 == 0 else COLOR_ROW_MAIN
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(
                color=COLOR_LINK if is_url else COLOR_TEXT_LIGHT,
                size=10, name="Segoe UI",
                underline="single" if is_url else None,
            )
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(header == "Overview"),
            )
            if is_url and val:
                cell.hyperlink = val

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions
    return ws


def write_summary_sheet(wb, stats):
    ws = wb.create_sheet(title="📊 Summary", index=0)
    ws.sheet_properties.tabColor = COLOR_HEADER_SUM
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "🎬 Cineby.gd Content Database"
    title_cell.font = Font(bold=True, size=18, color=COLOR_TEXT_LIGHT, name="Segoe UI")
    title_cell.fill = PatternFill("solid", fgColor=COLOR_DARK_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 45

    ws.merge_cells("A2:B2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Source: TMDB API + cineby.gd"
    sub_cell.font = Font(size=10, color=COLOR_TEXT_DIM, name="Segoe UI", italic=True)
    sub_cell.fill = PatternFill("solid", fgColor=COLOR_DARK_BG)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    rows = [
        ("Category",        "Total Items",   COLOR_HEADER_SUM),
        ("🎬 Movies",        stats.get("movies", 0),   None),
        ("📺 TV Shows",      stats.get("tv_shows", 0), None),
        ("🎌 Anime",         stats.get("anime", 0),    None),
        ("📡 Channels",      stats.get("channels", 0), None),
        ("━" * 20,           "━" * 10,                 None),
        ("📦 TOTAL CONTENT", stats.get("total", 0),    COLOR_HEADER_SUM),
    ]

    label_colors = {
        "🎬 Movies":        COLOR_HEADER_MOV,
        "📺 TV Shows":      COLOR_HEADER_TV,
        "🎌 Anime":         COLOR_HEADER_ANI,
        "📡 Channels":      COLOR_HEADER_CH,
        "📦 TOTAL CONTENT": COLOR_HEADER_SUM,
    }

    for r_idx, (label, value, force_color) in enumerate(rows, 4):
        ws.row_dimensions[r_idx].height = 26
        
        lc = ws.cell(row=r_idx, column=1, value=label)
        vc = ws.cell(row=r_idx, column=2, value=value)

        bg = force_color or COLOR_ROW_ALT if r_idx % 2 == 0 else COLOR_ROW_MAIN
        accent = label_colors.get(label)

        for cell in (lc, vc):
            cell.fill = PatternFill("solid", fgColor=accent or bg)
            cell.font = Font(
                bold=(accent is not None),
                size=12 if accent else 11,
                color=COLOR_TEXT_LIGHT,
                name="Segoe UI",
            )
            cell.alignment = Alignment(horizontal="left" if cell == lc else "right",
                                       vertical="center")

    # Add a note
    note_row = len(rows) + 6
    ws.merge_cells(f"A{note_row}:B{note_row}")
    note = ws.cell(row=note_row, column=1,
                   value="💡 Cineby uses TMDB IDs. Embed via Vidking: https://www.vidking.net/embed/movie/{tmdbId}")
    note.font = Font(size=9, color=COLOR_TEXT_DIM, name="Segoe UI", italic=True)
    note.fill = PatternFill("solid", fgColor=COLOR_DARK_BG)
    note.alignment = Alignment(vertical="center")
    ws.row_dimensions[note_row].height = 20

    for row in ws.iter_rows():
        for cell in row:
            if not cell.fill or not cell.fill.fgColor or cell.fill.fgColor.type == "none":
                cell.fill = PatternFill("solid", fgColor=COLOR_DARK_BG)

    return ws


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  🎬 Cineby.gd Content Scraper — Powered by TMDB API")
    print("=" * 65)
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), OUTPUT_FILE)
    mode_label = "INCREMENTAL (add new only)" if INCREMENTAL_UPDATE else "FULL (overwrite)"
    print(f"  Output file : {output_path}")
    print(f"  Run mode    : {mode_label}")
    print(f"  Max pages   : Movies={MAX_PAGES_MOVIES or 'ALL'}, "
          f"TV={MAX_PAGES_TV or 'ALL'}, Anime={MAX_PAGES_ANIME or 'ALL'}")
    print("=" * 65)

    client = TMDBClient()

    # ── Decide fetch strategy ──────────────────────────────────────────────────
    if INCREMENTAL_UPDATE and os.path.exists(output_path):
        print("\n📂  Loading existing data from Excel...")
        existing_movies, ex_movie_ids     = load_existing_rows(output_path, "🎬 Movies")
        existing_tv,     ex_tv_ids         = load_existing_rows(output_path, "📺 TV Shows")
        existing_anime,  ex_anime_ids      = load_existing_rows(output_path, "🎌 Anime (Series)")
        existing_animov, ex_animov_ids     = load_existing_rows(output_path, "🎌 Anime Movies")
        existing_ch,     ex_ch_ids         = load_existing_rows(output_path, "📡 Channels",
                                                                id_col="Network ID")
        print(f"  ✓ Loaded: {len(existing_movies):,} movies, {len(existing_tv):,} TV, "
              f"{len(existing_anime):,} anime series, {len(existing_animov):,} anime movies, "
              f"{len(existing_ch):,} channels")
        fetch_fn = fetch_incremental
    else:
        if INCREMENTAL_UPDATE:
            print("  (No existing file found — running full fetch)")
        existing_movies = existing_tv = existing_anime = existing_animov = existing_ch = []
        ex_movie_ids = ex_tv_ids = ex_anime_ids = ex_animov_ids = ex_ch_ids = set()
        fetch_fn = lambda client, ep, params, ex_ids, mp, desc: \
            client.fetch_all_pages(ep, params=params, max_pages=mp, desc=desc)

    # ── 1. MOVIES ─────────────────────────────────────────────────────────────────────────────
    print("\n🎬  Fetching Movies...")
    movie_lists = []

    # 1. Base lists
    endpoints = [
        ("/trending/movie/day", {"language": "en-US"}, "  Trending movies (today)"),
        ("/trending/movie/week", {"language": "en-US"}, "  Trending movies (week)"),
        ("/movie/popular", {"language": "en-US"}, "  Popular movies"),
        ("/movie/top_rated", {"language": "en-US"}, "  Top rated movie hall of fame"),
        ("/movie/now_playing", {"language": "en-US", "region": "US"}, "  Now playing (US)"),
        ("/movie/now_playing", {"language": "en-US", "region": "IN"}, "  Now playing (IN)"),
        ("/movie/upcoming", {"language": "en-US"}, "  Upcoming global"),
    ]

    # 2. Regional & Language Slicing (The secret to getting >10,000 movies)
    # TMDB caps every single query at 500 pages. To get more, we must split 
    # requests by region, language, or year.
    slices = [
        ("IN", "hi", "Hindi / Bollywood"),
        ("IN", "te", "Telugu / Tollywood"),
        ("IN", "ta", "Tamil / Kollywood"),
        ("KR", "ko", "Korean / K-Drama Movies"),
        ("JP", "ja", "Japanese / Anime-style"),
        ("ES", "es", "Spanish / Latin"),
        ("FR", "fr", "French"),
        ("US", "en", "Hollywood Mainstream"),
    ]
    for region_code, lang_code, label in slices:
        endpoints.append((
            "/discover/movie",
            {
                "sort_by": "popularity.desc",
                "watch_region": region_code,
                "with_original_language": lang_code,
                "language": "en-US"
            },
            f"  Slice: {label}"
        ))

    # 3. Yearly discovery for last 20 years (Deep archive)
    import datetime
    current_year = datetime.datetime.now().year
    for year in range(current_year, current_year - 21, -1):
        endpoints.append((
            "/discover/movie",
            {"sort_by": "popularity.desc", "primary_release_year": str(year), "language": "en-US"},
            f"  Deep Archive: {year}"
        ))

    for ep, params, desc in endpoints:
        # We fetch up to 100 pages for each specific slice/year.
        # This keeps the run time sane while providing massive coverage.
        # Total pages across all slices will be well over 3000.
        pages_to_fetch = 100 
        
        # For the global popular/trending, we can go higher (500)
        if "global" in desc.lower() or "popular movies" in desc.lower():
            pages_to_fetch = MAX_PAGES_MOVIES or 500
            
        items = fetch_fn(client, ep, params, ex_movie_ids, pages_to_fetch, desc)
        movie_lists.extend(items)

    new_movies  = dedupe(movie_lists)
    new_mov_data = [normalize_movie(m) for m in new_movies]
    # Merge: existing first, then new (existing rows keep their order)
    movies_data = existing_movies + [r for r in new_mov_data
                                     if r["TMDB ID"] not in ex_movie_ids]
    print(f"  → {len(new_mov_data):,} new movies fetched, "
          f"{len(movies_data):,} total in file")

    # ── 2. TV SHOWS ─────────────────────────────────────────────────────────────────────────────
    print("\n📺  Fetching TV Shows...")
    tv_lists = []

    tv_endpoints = [
        ("/trending/tv/day", {"language": "en-US"}, "  Trending TV (today)"),
        ("/trending/tv/week", {"language": "en-US"}, "  Trending TV (week)"),
        ("/tv/popular", {"language": "en-US"}, "  Popular TV series"),
        ("/tv/top_rated", {"language": "en-US"}, "  Top rated TV series"),
        ("/tv/on_the_air", {"language": "en-US"}, "  On the air"),
    ]

    # Regional TV Slicing
    for region_code, lang_code, label in slices:
        tv_endpoints.append((
            "/discover/tv",
            {
                "sort_by": "popularity.desc",
                "watch_region": region_code,
                "with_original_language": lang_code,
                "language": "en-US"
            },
            f"  TV Slice: {label}"
        ))

    for ep, params, desc in tv_endpoints:
        pages_to_fetch = 100
        if "global" in desc.lower() or "popular tv" in desc.lower():
            pages_to_fetch = MAX_PAGES_TV or 500
        items = fetch_fn(client, ep, params, ex_tv_ids, pages_to_fetch, desc)
        tv_lists.extend(items)

    new_tv      = dedupe(tv_lists)
    new_tv_data = [normalize_tv(t) for t in new_tv]
    tv_data     = existing_tv + [r for r in new_tv_data
                                  if r["TMDB ID"] not in ex_tv_ids]
    print(f"  → {len(new_tv_data):,} new TV shows fetched, "
          f"{len(tv_data):,} total in file")

    # ── 3. ANIME ─────────────────────────────────────────────────────────────────────────────
    print("\n🎌  Fetching Anime...")
    anime_lists  = []
    animov_lists = []

    anime_endpoints = [
        # Anime = Animation genre (16) + Japanese origin
        ("/discover/tv", {
            "sort_by": "popularity.desc",
            "with_genres": "16",
            "with_origin_country": "JP",
            "language": "en-US",
        }, "  Anime (JP Animation)"),
        # Also fetch by keyword "anime"
        ("/discover/tv", {
            "sort_by": "popularity.desc",
            "with_keywords": "210024",  # TMDB keyword ID for "anime"
            "language": "en-US",
        }, "  Anime (keyword)"),
        # Korean animation
        ("/discover/tv", {
            "sort_by": "popularity.desc",
            "with_genres": "16",
            "with_origin_country": "KR",
            "language": "en-US",
        }, "  Anime (KR Animation)"),
    ]

    for ep, params, desc in anime_endpoints:
        items = fetch_fn(client, ep, params, ex_anime_ids, MAX_PAGES_ANIME, desc)
        anime_lists.extend(items)

    anime_movie_items = fetch_fn(
        client,
        "/discover/movie",
        {"with_genres": "16", "with_origin_country": "JP",
         "sort_by": "popularity.desc", "language": "en-US"},
        ex_animov_ids, 50,
        "  Anime Movies (JP)"
    )

    new_anime_tv     = dedupe(anime_lists)
    new_anime_tv_data = [normalize_tv(a, is_anime=True) for a in new_anime_tv]
    all_anime_data   = existing_anime + [r for r in new_anime_tv_data
                                         if r["TMDB ID"] not in ex_anime_ids]

    new_anime_movies  = dedupe(anime_movie_items)
    new_anime_mov_data = [normalize_movie(m) for m in new_anime_movies]
    anime_movie_data  = existing_animov + [r for r in new_anime_mov_data
                                            if r["TMDB ID"] not in ex_animov_ids]

    print(f"  → {len(new_anime_tv_data):,} new anime series fetched, "
          f"{len(all_anime_data):,} total")
    print(f"  → {len(new_anime_mov_data):,} new anime movies fetched, "
          f"{len(anime_movie_data):,} total")

    # ── 4. CHANNELS / NETWORKS ────────────────────────────────────────────────────────────────
    new_channels  = fetch_networks(client, max_pages=MAX_PAGES_CHANNELS)
    new_ch_ids    = {c["Network ID"] for c in new_channels}
    channels_data = existing_ch + [c for c in new_channels
                                    if c["Network ID"] not in ex_ch_ids]
    print(f"  → {len(new_channels):,} channels fetched, "
          f"{len(channels_data):,} total in file")

    # ── 5. BUILD EXCEL ─────────────────────────────────────────────────────────────────────────────
    print(f"\n📝  Building Excel file...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    stats = {
        "movies":   len(movies_data),
        "tv_shows": len(tv_data),
        "anime":    len(all_anime_data) + len(anime_movie_data),
        "channels": len(channels_data),
        "total":    len(movies_data) + len(tv_data) + len(all_anime_data)
                    + len(channels_data),
    }

    # Summary sheet (index 0)
    write_summary_sheet(wb, stats)

    # Movies sheet
    movie_headers = [
        "TMDB ID", "Title", "Overview", "Release Date",
        "Rating (TMDB)", "Vote Count", "Popularity", "Language",
        "Genres", "Cineby URL", "Vidking Embed", "Poster", "Adult",
    ]
    write_sheet(wb, "🎬 Movies", COLOR_HEADER_MOV, COLOR_HEADER_MOV,
                movies_data, movie_headers,
                url_headers={"Cineby URL", "Vidking Embed", "Poster"})
    print(f"  ✓ Movies sheet ({len(movies_data):,} rows)")

    # TV Shows sheet
    tv_headers = [
        "TMDB ID", "Title", "Overview", "First Air Date",
        "Rating (TMDB)", "Vote Count", "Popularity", "Language",
        "Genres", "Origin Country", "Cineby URL", "Cineby Ep1 URL",
        "Vidking Embed", "Poster",
    ]
    write_sheet(wb, "📺 TV Shows", COLOR_HEADER_TV, COLOR_HEADER_TV,
                tv_data, tv_headers,
                url_headers={"Cineby URL", "Cineby Ep1 URL", "Vidking Embed", "Poster"})
    print(f"  ✓ TV Shows sheet ({len(tv_data):,} rows)")

    # Anime (TV Series) sheet
    write_sheet(wb, "🎌 Anime (Series)", COLOR_HEADER_ANI, COLOR_HEADER_ANI,
                all_anime_data, tv_headers,
                url_headers={"Cineby URL", "Cineby Ep1 URL", "Vidking Embed", "Poster"})
    print(f"  ✓ Anime Series sheet ({len(all_anime_data):,} rows)")

    # Anime Movies sheet
    write_sheet(wb, "🎌 Anime Movies", COLOR_HEADER_ANI, COLOR_HEADER_ANI,
                anime_movie_data, movie_headers,
                url_headers={"Cineby URL", "Vidking Embed", "Poster"})
    print(f"  ✓ Anime Movies sheet ({len(anime_movie_data):,} rows)")

    # Channels sheet
    channel_headers = [
        "Network ID", "Name", "Country", "Headquarters",
        "Homepage", "TMDB Page", "Logo",
    ]
    write_sheet(wb, "📡 Channels", COLOR_HEADER_CH, COLOR_HEADER_CH,
                channels_data, channel_headers,
                url_headers={"Homepage", "TMDB Page", "Logo"})
    print(f"  ✓ Channels sheet ({len(channels_data):,} rows)")

    # Save
    wb.save(output_path)
    file_size_mb = os.path.getsize(output_path) / (1024 * 1024)

    print("\n" + "=" * 65)
    print(f"  ✅  Done! Saved to: {output_path}")
    print(f"      File size  : {file_size_mb:.1f} MB")
    print(f"      Run mode   : {mode_label}")
    print(f"      Movies     : {stats['movies']:,}")
    print(f"      TV Shows   : {stats['tv_shows']:,}")
    print(f"      Anime      : {stats['anime']:,}")
    print(f"      Channels   : {stats['channels']:,}")
    print(f"      TOTAL      : {stats['total']:,} items")
    print("=" * 65)


if __name__ == "__main__":
    main()
