"""
==========================================================================
  CinebyHub — Master Runner  (Smart Scheduler Edition)
==========================================================================
  Runs the full pipeline in a loop, every 12 hours:

    STEP 1 → cineby_scraper.py       (fetch TMDB content → Excel)
    STEP 2 → linkvertise_api_lite.py (wrap NEW URLs → Linkvertise links)
              ↑ Only runs when new rows were actually added by the scraper.
    STEP 3 → npm run dev             (Vite web app — stays alive always)
              ↑ Launched once on startup, stays running in background.

  Every 12 hours the scheduler wakes up, re-runs the scraper, checks
  if any new rows were added, and only then calls the LV generator.

  Row-count state is stored in:
      public/_checkpoints/row_counts.json

  Usage:
    python run_all.py                ← full loop (default)
    python run_all.py --once         ← run pipeline once, no loop
    python run_all.py --skip-lv      ← never run Linkvertise step
    python run_all.py --no-scrape    ← skip TMDB scrape
    python run_all.py --web-only     ← just launch the web server
    python run_all.py --interval 6   ← check every 6 hours (default: 12)
    python run_all.py --no-browser   ← don't auto-open browser tab
==========================================================================
"""

import sys
import os
import json
import subprocess
import argparse
import webbrowser
import time
import threading
from pathlib import Path
from datetime import datetime, timedelta

BASE            = Path(__file__).parent
SOURCE_EXCEL    = BASE / "cineby_content.xlsx"
CHECKPOINT_DIR  = BASE / "public" / "_checkpoints"
ROW_COUNT_FILE  = CHECKPOINT_DIR / "row_counts.json"

CHECK_INTERVAL_HOURS = 12   # default; overridable via --interval

# ── Colour helpers ────────────────────────────────────────────────────────────
GREEN  = "\033[92m"
YELLOW = "\033[93m"
RED    = "\033[91m"
CYAN   = "\033[96m"
MAGENTA= "\033[95m"
BOLD   = "\033[1m"
DIM    = "\033[2m"
RESET  = "\033[0m"

def banner(text, color=CYAN):
    width = 65
    print(f"\n{color}{BOLD}{'═' * width}")
    print(f"  {text}")
    print(f"{'═' * width}{RESET}\n")

def step_header(n, total, title):
    print(f"\n{BOLD}{CYAN}[{n}/{total}] {title}{RESET}")
    print(f"{'─' * 60}")

def success(msg): print(f"{GREEN}  ✅  {msg}{RESET}", flush=True)
def warn(msg):    print(f"{YELLOW}  ⚠   {msg}{RESET}", flush=True)
def error(msg):   print(f"{RED}  ✗   {msg}{RESET}", flush=True)
def info(msg):    print(f"      {msg}", flush=True)
def dim(msg):     print(f"{DIM}  {msg}{RESET}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
#  Row-count state  (stored in public/_checkpoints/row_counts.json)
# ─────────────────────────────────────────────────────────────────────────────

def _load_row_counts() -> dict:
    """Load previously saved row counts per sheet."""
    if ROW_COUNT_FILE.exists():
        try:
            return json.loads(ROW_COUNT_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_row_counts(counts: dict):
    CHECKPOINT_DIR.mkdir(parents=True, exist_ok=True)
    ROW_COUNT_FILE.write_text(
        json.dumps(counts, indent=2, ensure_ascii=False),
        encoding="utf-8"
    )


def _count_excel_rows(path: Path) -> dict:
    """
    Return {sheet_name: row_count} for all sheets in the Excel file.
    Uses openpyxl in read-only mode so it's fast even on large files.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        counts = {}
        for name in wb.sheetnames:
            ws = wb[name]
            # max_row includes the header, subtract 1 for data rows
            counts[name] = max(0, (ws.max_row or 1) - 1)
        wb.close()
        return counts
    except Exception as e:
        warn(f"Could not read row counts from Excel: {e}")
        return {}


def detect_new_rows(path: Path) -> tuple[bool, dict, dict]:
    """
    Compare current row counts with the saved snapshot.

    Returns:
        (has_new_rows, old_counts, new_counts)
    """
    old_counts = _load_row_counts()
    new_counts = _count_excel_rows(path)

    if not old_counts:
        # First run — treat as "new data" so LV runs at least once
        info("No previous row snapshot found → treating as new data.")
        return True, old_counts, new_counts

    has_new = False
    for sheet, count in new_counts.items():
        prev = old_counts.get(sheet, 0)
        if count > prev:
            diff = count - prev
            info(f"  📈  {sheet}: {prev:,} → {count:,}  (+{diff:,} new rows)")
            has_new = True
        else:
            dim(f"  {sheet}: {count:,} rows (no change)")

    return has_new, old_counts, new_counts


# ─────────────────────────────────────────────────────────────────────────────
#  Script runner
# ─────────────────────────────────────────────────────────────────────────────

def run_script(script_name: str, description: str) -> bool:
    """Run a Python script and stream its output live."""
    script_path = BASE / script_name
    if not script_path.exists():
        error(f"Script not found: {script_path}")
        return False

    info(f"Running: {script_path}")
    try:
        result = subprocess.run(
            [sys.executable, "-u", str(script_path)],
            cwd=str(BASE),
            env={**os.environ, "PYTHONUNBUFFERED": "1"},
        )
        if result.returncode != 0:
            error(f"{description} exited with code {result.returncode}")
            return False
        success(f"{description} completed successfully.")
        return True
    except KeyboardInterrupt:
        warn(f"{description} interrupted by user.")
        return False
    except Exception as e:
        error(f"{description} failed: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
#  Web app (runs persistently in background)
# ─────────────────────────────────────────────────────────────────────────────

_webapp_proc: subprocess.Popen | None = None

def start_webapp(open_browser: bool = True):
    """Launch Vite dev server in background (non-blocking)."""
    global _webapp_proc
    if _webapp_proc and _webapp_proc.poll() is None:
        dim("Web server is already running.")
        return

    info("Starting Vite dev server (background)…")
    try:
        _webapp_proc = subprocess.Popen(
            ["npm", "run", "dev"],
            # ["npm", "run", "preview", "--", "--host", "0.0.0.0", "--port", os.environ.get("PORT", "8080")],
            cwd=str(BASE),
            shell=True,      # required on Windows for npm
            env={**os.environ},
        )
        info(f"Web server PID: {_webapp_proc.pid}")

        if open_browser:
            def _open():
                time.sleep(4)
                webbrowser.open("http://localhost:5173")
            threading.Thread(target=_open, daemon=True).start()

        success("Web server started → http://localhost:5173")
    except Exception as e:
        error(f"Failed to start web server: {e}")


def stop_webapp():
    global _webapp_proc
    if _webapp_proc and _webapp_proc.poll() is None:
        _webapp_proc.terminate()
        try:
            _webapp_proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            _webapp_proc.kill()
        _webapp_proc = None
        dim("Web server stopped.")


# ─────────────────────────────────────────────────────────────────────────────
#  One pipeline cycle
# ─────────────────────────────────────────────────────────────────────────────

def run_pipeline_cycle(args) -> None:
    """
    Run one full scrape + (optional) LV generation cycle.

    Logic:
      1. Run TMDB scraper (unless --no-scrape / --web-only)
      2. Count rows in the updated Excel
      3. If new rows found (or first run) → run Linkvertise generator
      4. Save the new row counts as the baseline for next cycle
    """
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    banner(f"CinebyHub — Pipeline Cycle  [{now_str}]")

    step = 1
    total_steps = 2  # scrape + (maybe) LV

    # ── Step 1: TMDB Scraper ─────────────────────────────────────────────────
    if not args.no_scrape and not args.web_only and not args.skip_lv_only:
        step_header(step, total_steps, "TMDB Scraper")
        step += 1
        scrape_ok = run_script("cineby_scraper.py", "TMDB Scraper")
        if not scrape_ok:
            warn("Scraper failed. Skipping Linkvertise step this cycle.")
            return
    else:
        info("Scraper step skipped.")
        step += 1

    # ── Check for new rows ────────────────────────────────────────────────────
    if args.skip_lv or args.web_only:
        info("Linkvertise step disabled by flag.")
        return

    if not SOURCE_EXCEL.exists():
        warn(f"Source Excel not found: {SOURCE_EXCEL}")
        return

    print(f"\n{CYAN}  🔍  Checking for new content rows…{RESET}")
    has_new, old_counts, new_counts = detect_new_rows(SOURCE_EXCEL)

    # ── Step 2: Linkvertise Generator ────────────────────────────────────────
    step_header(step, total_steps, "Linkvertise Link Generator")

    if has_new or args.force_lv:
        if args.force_lv and not has_new:
            info("--force-lv flag set — running Linkvertise even with no new rows.")
        else:
            total_new_rows = sum(
                max(0, new_counts.get(s, 0) - old_counts.get(s, 0))
                for s in new_counts
            )
            info(f"  🆕  {total_new_rows:,} new rows detected → running Linkvertise generator…")

        lv_ok = run_script("linkvertise_api_lite.py", "Linkvertise Link Generator")

        # Save updated counts regardless of LV success
        _save_row_counts(new_counts)

        if lv_ok:
            success("All new links generated successfully.")
        else:
            warn("Linkvertise generator had errors. Row counts saved anyway.")
    else:
        success("No new content rows found — skipping Linkvertise generator. ✓")
        # Still update the baseline (counts haven't changed, but timestamp matters)
        _save_row_counts(new_counts)


# ─────────────────────────────────────────────────────────────────────────────
#  Countdown display
# ─────────────────────────────────────────────────────────────────────────────

def _countdown(wake_at: datetime):
    """Show a live countdown until next scheduled run (updates every minute)."""
    try:
        while True:
            remaining = wake_at - datetime.now()
            if remaining.total_seconds() <= 0:
                break
            h, rem = divmod(int(remaining.total_seconds()), 3600)
            m = rem // 60
            print(
                f"\r{DIM}  ⏰  Next check in {h}h {m:02d}m  "
                f"(at {wake_at.strftime('%H:%M:%S')})   {RESET}",
                end="", flush=True
            )
            time.sleep(30)
        print()  # newline
    except KeyboardInterrupt:
        raise


# ─────────────────────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="CinebyHub — Smart pipeline runner with 12-hour scheduling",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python run_all.py                  Full loop (scrape → LV if new → web)
  python run_all.py --once           Run pipeline once, then exit
  python run_all.py --skip-lv        Loop without Linkvertise step
  python run_all.py --no-scrape      Skip TMDB scrape each cycle
  python run_all.py --web-only       Just launch the web server
  python run_all.py --force-lv       Always run LV even without new rows
  python run_all.py --interval 6     Check every 6 hours (default: 12)
  python run_all.py --no-browser     Don't auto-open browser tab
        """
    )
    parser.add_argument("--once",        action="store_true", help="Run once and exit (no loop)")
    parser.add_argument("--skip-lv",     action="store_true", help="Never run Linkvertise step")
    parser.add_argument("--skip-lv-only",action="store_true", dest="skip_lv_only",
                        help="Skip only the scraper (run LV + web)")
    parser.add_argument("--no-scrape",   action="store_true", help="Skip TMDB scraper each cycle")
    parser.add_argument("--web-only",    action="store_true", help="Only start the web server")
    parser.add_argument("--force-lv",    action="store_true", help="Always run LV regardless of new rows")
    parser.add_argument("--no-browser",  action="store_true", help="Don't auto-open browser tab")
    parser.add_argument("--interval",    type=float, default=CHECK_INTERVAL_HOURS,
                        metavar="HOURS",  help=f"Hours between checks (default: {CHECK_INTERVAL_HOURS})")
    args = parser.parse_args()

    banner("CinebyHub — Master Pipeline Runner")

    # ── Always start the web server first (runs in background) ───────────────
    if not args.once:
        info(f"Scheduler interval : every {args.interval}h")
        info(f"Auto-open browser  : {'no' if args.no_browser else 'yes'}")
        info(f"Scraper            : {'skipped' if args.no_scrape or args.web_only else 'enabled'}")
        info(f"Linkvertise        : {'skipped' if args.skip_lv or args.web_only else 'smart (runs only when new rows detected)'}")
        print()

    if not args.once:
        start_webapp(open_browser=not args.no_browser)

    if args.web_only:
        info("--web-only mode. Press Ctrl+C to stop.")
        try:
            while True:
                time.sleep(60)
        except KeyboardInterrupt:
            print()
            warn("Shutting down…")
            stop_webapp()
        return

    # ── Pipeline loop ─────────────────────────────────────────────────────────
    run_count = 0
    try:
        while True:
            run_count += 1
            if run_count > 1:
                print()
                info(f"{'─' * 60}")
                info(f"  Scheduled wake-up #{run_count}")
                info(f"{'─' * 60}")

            run_pipeline_cycle(args)

            if args.once:
                break

            # Schedule next run
            interval_sec = args.interval * 3600
            wake_at = datetime.now() + timedelta(seconds=interval_sec)
            print(f"\n{CYAN}  🕐  Pipeline complete. Sleeping {args.interval}h until next check…{RESET}")
            info(f"  Next run at: {wake_at.strftime('%Y-%m-%d %H:%M:%S')}")
            info("  (Ctrl+C to stop the scheduler and web server)\n")

            _countdown(wake_at)

            # Sleep in small chunks so Ctrl+C is responsive
            sleep_end = time.time() + interval_sec
            while time.time() < sleep_end:
                time.sleep(1)

    except KeyboardInterrupt:
        print(f"\n\n{YELLOW}  Interrupted by user.{RESET}")
    finally:
        stop_webapp()
        banner("CinebyHub scheduler stopped.", YELLOW)


if __name__ == "__main__":
    main()
