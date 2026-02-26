"""
==========================================================================
  Cineby × Linkvertise Link Generator
==========================================================================
  Reads cineby_content.xlsx (from cineby_scraper.py) sheet by sheet,
  wraps the correct URL through Linkvertise, and writes the results
  back to public/cineby_content.xlsx.

  Uses the same 'linkvertise' Python package (no API key needed).

  URL strategy per sheet:
    🎬 Movies          → Vidking Embed
    📺 TV Shows        → Vidking Embed
    🎌 Anime (Series)  → Vidking Embed
    🎌 Anime Movies    → Vidking Embed
    📡 Channels        → Homepage

  Re-running is SAFE — already-processed rows are skipped.
  Progress is saved to a lightweight CSV checkpoint file every N rows,
  and the heavy formatted Excel is written only once per sheet.
==========================================================================
"""

import sys
import os
import shutil
import time
from pathlib import Path

# Force unbuffered output so everything appears in real-time
os.environ.setdefault("PYTHONUNBUFFERED", "1")

def log(msg="", **kwargs):
    print(msg, flush=True, **kwargs)

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────────────────────────────────────
USER_ID      = "738317"

BASE_DIR     = Path(__file__).parent
SOURCE_EXCEL = BASE_DIR / "cineby_content.xlsx"
OUTPUT_EXCEL = BASE_DIR / "public" / "cineby_content.xlsx"
CHECKPOINT_DIR = BASE_DIR / "public" / "_checkpoints"

# Save a lightweight CSV checkpoint every N successfully processed rows.
# The heavy Excel is only written once per completed sheet.
SAVE_EVERY   = 500

SHEET_CONFIG = {
    "🎬 Movies":          "Vidking Embed",
    "📺 TV Shows":        "Vidking Embed",
    "🎌 Anime (Series)":  "Vidking Embed",
    "🎌 Anime Movies":    "Vidking Embed",
    "📡 Channels":        "Homepage",
}
# ─────────────────────────────────────────────────────────────────────────────


def check_deps():
    missing = []
    for pkg in ["linkvertise", "pandas", "openpyxl"]:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pkg)
    if missing:
        log(f"[!] Missing packages — install with:")
        log(f"    pip install {' '.join(missing)}")
        sys.exit(1)

check_deps()

import pandas as pd
from linkvertise import LinkvertiseClient


def _checkpoint_path(sheet_name: str) -> Path:
    """Return the CSV checkpoint file path for a given sheet."""
    safe = sheet_name.replace("/", "_").replace("\\", "_").replace(":", "_")
    return CHECKPOINT_DIR / f"{safe}.csv"


def _load_checkpoint(sheet_name: str) -> dict:
    """
    Load a saved checkpoint CSV → dict { original_index: lv_url }.
    Returns empty dict if no checkpoint exists.
    """
    cp = _checkpoint_path(sheet_name)
    if not cp.exists():
        return {}
    try:
        cdf = pd.read_csv(str(cp), dtype=str).fillna("")
        return dict(zip(cdf["idx"].astype(str), cdf["lv_url"]))
    except Exception:
        return {}


def _save_checkpoint(sheet_name: str, cp_data: dict):
    """Persist the checkpoint dict to a small CSV."""
    CHECKPOINT_DIR.mkdir(parents=True, exist_ok=True)
    cp = _checkpoint_path(sheet_name)
    rows = [{"idx": k, "lv_url": v} for k, v in cp_data.items()]
    pd.DataFrame(rows).to_csv(str(cp), index=False)


def main():
    log("=" * 65)
    log("  🔗 Cineby × Linkvertise Link Generator")
    log("=" * 65)
    log(f"  Source : {SOURCE_EXCEL}")
    log(f"  Output : {OUTPUT_EXCEL}")
    log(f"  UserID : {USER_ID}")
    log("=" * 65)

    # ── Validate ──────────────────────────────────────────────────────────────
    if not SOURCE_EXCEL.exists():
        log(f"\n[✗] Source not found: {SOURCE_EXCEL}")
        log("    Run cineby_scraper.py first.")
        sys.exit(1)

    OUTPUT_EXCEL.parent.mkdir(parents=True, exist_ok=True)
    CHECKPOINT_DIR.mkdir(parents=True, exist_ok=True)

    # Copy source → output on first run
    if not OUTPUT_EXCEL.exists():
        log("\n📄  First run — copying source to output...")
        shutil.copy2(SOURCE_EXCEL, OUTPUT_EXCEL)
        log("    Done.")
    else:
        log("\n📂  Resuming from existing output file (incremental mode)...")

    # ── Linkvertise client ────────────────────────────────────────────────────
    client = LinkvertiseClient()
    log("[✓] Linkvertise client ready")

    total_new = 0

    for sheet_name, url_col in SHEET_CONFIG.items():
        log(f"\n── {sheet_name} {'─' * (50 - len(sheet_name))}")
        log(f"  Loading sheet from SOURCE...", end=" ")
        sys.stdout.flush()

        # Always load from the SOURCE excel (clean data, no partial LV columns)
        try:
            df = pd.read_excel(str(SOURCE_EXCEL), sheet_name=sheet_name, dtype=str)
            df = df.fillna("")
        except Exception as e:
            log(f"SKIP — {e}")
            continue

        log(f"OK ({len(df):,} rows)")

        if url_col not in df.columns:
            log(f"  [!] Column '{url_col}' not found — skipping.")
            continue

        # Ensure LV column exists
        if "Linkvertise_Link" not in df.columns:
            df["Linkvertise_Link"] = ""

        # ── Load checkpoint (fast, from CSV) ──────────────────────────────────
        cp_data = _load_checkpoint(sheet_name)

        # Also try to pull already-done links from the OUTPUT Excel
        # (covers the case where output Excel was written but checkpoint deleted)
        if not cp_data and OUTPUT_EXCEL.exists():
            try:
                odf = pd.read_excel(str(OUTPUT_EXCEL), sheet_name=sheet_name, dtype=str).fillna("")
                if "Linkvertise_Link" in odf.columns:
                    for i, lv in enumerate(odf["Linkvertise_Link"]):
                        if lv.startswith("http"):
                            cp_data[str(i)] = lv
                    if cp_data:
                        log(f"  ↩  Recovered {len(cp_data):,} links from output Excel into checkpoint.")
                        _save_checkpoint(sheet_name, cp_data)
            except Exception:
                pass

        # Merge checkpoint back into df
        for idx_str, lv_url in cp_data.items():
            try:
                df.at[int(idx_str), "Linkvertise_Link"] = lv_url
            except Exception:
                pass

        # Determine what still needs processing
        needs_lv_mask = (df["Linkvertise_Link"] == "") & df[url_col].str.startswith("http", na=False)
        needs_lv = df[needs_lv_mask]
        n_done = int((df["Linkvertise_Link"] != "").sum())
        n_todo = len(needs_lv)

        log(f"  Already processed : {n_done:,}")
        log(f"  Rows to process   : {n_todo:,}")

        if n_todo == 0:
            log("  ✓ All rows already have Linkvertise links!")
            # Still write the Excel sheet if this is the first time we're "done"
            _write_excel_sheet(df, sheet_name, OUTPUT_EXCEL)
            log(f"  ✅  Sheet '{sheet_name}' is complete in output Excel.")
            continue

        success = 0
        errors  = 0
        t0 = time.time()

        for pos, (idx, row) in enumerate(needs_lv.iterrows(), 1):
            target_url = row[url_col]

            # Progress line (overwrite in place)
            elapsed = time.time() - t0
            rate = pos / elapsed if elapsed > 0 else 0
            eta = (n_todo - pos) / rate if rate > 0 else 0
            print(
                f"\r  [{pos}/{n_todo}] success={success} errors={errors}"
                f"  {rate:.1f} links/s  ETA {eta/60:.1f}min   ",
                end="", flush=True
            )

            try:
                lv_url = client.linkvertise(USER_ID, target_url)
                df.at[idx, "Linkvertise_Link"] = lv_url
                cp_data[str(idx)] = lv_url
                success += 1
                total_new += 1
            except Exception as e:
                errors += 1
                print(f"\n  [!] Row {idx}: {str(e)[:80]}", flush=True)

            # Periodic lightweight checkpoint save (very fast — just CSV)
            if success > 0 and success % SAVE_EVERY == 0:
                print(f"\n  💾 Checkpoint saved ({n_done + success:,} total links)...", flush=True)
                _save_checkpoint(sheet_name, cp_data)

        print(flush=True)  # newline after the \r progress

        # Save final checkpoint
        _save_checkpoint(sheet_name, cp_data)

        # Heavy Excel write — only ONCE per sheet completion
        log(f"  � Writing formatted Excel sheet '{sheet_name}'...")
        _write_excel_sheet(df, sheet_name, OUTPUT_EXCEL)
        log(f"  ✅  Created {success:,} new LV links ({errors} errors)")

    log(f"\n{'=' * 65}")
    log(f"  ✅  Done!  Total new LV links created: {total_new:,}")
    log(f"  Output: {OUTPUT_EXCEL}")
    log(f"{'=' * 65}")


def _write_excel_sheet(df: "pd.DataFrame", sheet_name: str, output_path: Path):
    """
    Write ONE sheet back into the Excel file without touching other sheets.
    Uses openpyxl in append mode.

    Called ONCE per sheet (not every N rows), so heavy formatting is fine.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    WIDTHS = {
        "TMDB ID": 10, "Network ID": 10, "Title": 36, "Name": 36,
        "Overview": 50, "Release Date": 13, "First Air Date": 13,
        "Rating (TMDB)": 13, "Vote Count": 11, "Popularity": 11,
        "Language": 9, "Genres": 28, "Origin Country": 14,
        "Poster": 14, "Backdrop": 14, "Cineby URL": 36,
        "Cineby Ep1 URL": 38, "Vidking Embed": 42,
        "Homepage": 38, "TMDB Page": 38, "Country": 9,
        "Headquarters": 22, "Logo": 14, "Linkvertise_Link": 48,
        "Adult": 8, "Is Anime": 9,
    }
    TAB_COLORS = {
        "🎬 Movies":          "dc2626",
        "📺 TV Shows":        "0ea5e9",
        "🎌 Anime (Series)":  "7c3aed",
        "🎌 Anime Movies":    "9333ea",
        "📡 Channels":        "059669",
    }

    wb = openpyxl.load_workbook(str(output_path))

    # Remove old version of this sheet
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = TAB_COLORS.get(sheet_name, "334155")

    headers = list(df.columns)
    lv_idx_1 = headers.index("Linkvertise_Link") + 1 if "Linkvertise_Link" in headers else -1

    HDR_COLOR = TAB_COLORS.get(sheet_name, "1e293b")
    hdr_fill  = PatternFill("solid", fgColor=HDR_COLOR)

    # Header row
    ws.row_dimensions[1].height = 28
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Segoe UI")
        c.alignment = Alignment(horizontal="center", vertical="center")

    EVEN_FILL = PatternFill("solid", fgColor="111827")
    ODD_FILL  = PatternFill("solid", fgColor="0d1117")
    LV_FILL   = PatternFill("solid", fgColor="064e3b")

    # Data rows
    for ri, row_tuple in enumerate(df.itertuples(index=False), 2):
        ws.row_dimensions[ri].height = 18
        base_fill = EVEN_FILL if ri % 2 == 0 else ODD_FILL
        for ci, val in enumerate(row_tuple, 1):
            is_lv = (ci == lv_idx_1)
            val_str = "" if val is None or (isinstance(val, float) and val != val) else str(val)
            c = ws.cell(row=ri, column=ci, value=val_str)
            c.fill = LV_FILL if is_lv else base_fill
            c.font = Font(
                size=9, name="Segoe UI",
                color="10b981" if is_lv else "e2e8f0",
                underline="single" if val_str.startswith("http") else None,
            )
            c.alignment = Alignment(vertical="center")
            if val_str.startswith("http"):
                c.hyperlink = val_str

    # Column widths
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS.get(h, 16)

    ws.auto_filter.ref = ws.dimensions

    wb.save(str(output_path))
    wb.close()


if __name__ == "__main__":
    main()
